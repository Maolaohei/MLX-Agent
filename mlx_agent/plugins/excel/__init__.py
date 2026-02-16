"""
Excel插件 - Excel表格处理
迁移自 OpenClaw skill: excel
"""
import asyncio
import json
import os
from pathlib import Path
from typing import Any, Dict, List

from ..base import Plugin, register_plugin


@register_plugin
class ExcelPlugin(Plugin):
    """Excel表格处理插件"""
    
    @property
    def name(self) -> str:
        return "excel"
    
    @property
    def description(self) -> str:
        return "Excel文件读写、编辑、格式化、导出CSV/JSON/Markdown"
    
    async def _setup(self):
        """初始化插件"""
        self.workspace = self._config.get("workspace", "./workspace/excel")
        os.makedirs(self.workspace, exist_ok=True)
        
        # 检查依赖
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            proc = await asyncio.create_subprocess_exec(
                "pip", "install", "openpyxl", "-q",
                stdout=asyncio.subprocess.DEVNULL,
                stderr=asyncio.subprocess.DEVNULL
            )
            await proc.communicate()
    
    def get_tools(self) -> List[Dict]:
        """返回插件提供的工具"""
        return [
            {
                "type": "function",
                "function": {
                    "name": "excel_info",
                    "description": "获取Excel文件信息（工作表列表、行列数等）",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "file_path": {
                                "type": "string",
                                "description": "Excel文件路径"
                            }
                        },
                        "required": ["file_path"]
                    }
                }
            },
            {
                "type": "function",
                "function": {
                    "name": "excel_read",
                    "description": "读取Excel工作表内容",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "file_path": {
                                "type": "string",
                                "description": "Excel文件路径"
                            },
                            "sheet": {
                                "type": "string",
                                "description": "工作表名称，默认第一个"
                            },
                            "range": {
                                "type": "string",
                                "description": "单元格范围，如 A1:D10"
                            },
                            "format": {
                                "type": "string",
                                "enum": ["json", "markdown", "csv"],
                                "description": "输出格式",
                                "default": "json"
                            },
                            "has_header": {
                                "type": "boolean",
                                "description": "第一行是否为表头",
                                "default": True
                            }
                        },
                        "required": ["file_path"]
                    }
                }
            },
            {
                "type": "function",
                "function": {
                    "name": "excel_read_cell",
                    "description": "读取指定单元格内容",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "file_path": {
                                "type": "string",
                                "description": "Excel文件路径"
                            },
                            "cell": {
                                "type": "string",
                                "description": "单元格地址，如 B5"
                            },
                            "sheet": {
                                "type": "string",
                                "description": "工作表名称"
                            }
                        },
                        "required": ["file_path", "cell"]
                    }
                }
            },
            {
                "type": "function",
                "function": {
                    "name": "excel_create",
                    "description": "创建新Excel工作簿",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "output": {
                                "type": "string",
                                "description": "输出文件路径"
                            },
                            "sheets": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": "工作表名称列表"
                            }
                        },
                        "required": ["output"]
                    }
                }
            },
            {
                "type": "function",
                "function": {
                    "name": "excel_write",
                    "description": "向Excel写入数据",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "file_path": {
                                "type": "string",
                                "description": "Excel文件路径"
                            },
                            "sheet": {
                                "type": "string",
                                "description": "工作表名称"
                            },
                            "data": {
                                "type": "array",
                                "items": {"type": "array"},
                                "description": "二维数组数据"
                            },
                            "start_cell": {
                                "type": "string",
                                "description": "起始单元格，默认A1"
                            }
                        },
                        "required": ["file_path", "data"]
                    }
                }
            },
            {
                "type": "function",
                "function": {
                    "name": "excel_export",
                    "description": "导出Excel为其他格式",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "file_path": {
                                "type": "string",
                                "description": "Excel文件路径"
                            },
                            "output": {
                                "type": "string",
                                "description": "输出文件路径"
                            },
                            "format": {
                                "type": "string",
                                "enum": ["csv", "json", "markdown"],
                                "description": "导出格式"
                            },
                            "sheet": {
                                "type": "string",
                                "description": "要导出的工作表"
                            }
                        },
                        "required": ["file_path", "output", "format"]
                    }
                }
            }
        ]
    
    async def handle_tool(self, tool_name: str, params: Dict[str, Any]) -> Dict[str, Any]:
        """处理工具调用"""
        handlers = {
            "excel_info": self._info,
            "excel_read": self._read,
            "excel_read_cell": self._read_cell,
            "excel_create": self._create,
            "excel_write": self._write,
            "excel_export": self._export
        }
        
        if tool_name not in handlers:
            return {"success": False, "error": f"Unknown tool: {tool_name}"}
        
        try:
            return await handlers[tool_name](params)
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def _info(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """获取文件信息"""
        file_path = params["file_path"]
        
        if not os.path.exists(file_path):
            return {"success": False, "error": f"文件不存在: {file_path}"}
        
        cmd = [
            "python3", "-c",
            f"""
from openpyxl import load_workbook
import json

wb = load_workbook("{file_path}", read_only=True, data_only=True)
sheets = []

for name in wb.sheetnames:
    ws = wb[name]
    sheets.append({{
        "name": name,
        "rows": ws.max_row,
        "columns": ws.max_column,
        "dimensions": ws.dimensions
    }})

info = {{
    "sheets": sheets,
    "sheet_count": len(sheets),
    "file_size": os.path.getsize("{file_path}")
}}
print(json.dumps(info, ensure_ascii=False))
            """
        ]
        
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )
        stdout, stderr = await proc.communicate()
        
        if proc.returncode != 0:
            return {"success": False, "error": stderr.decode()}
        
        info = json.loads(stdout.decode())
        return {"success": True, **info}
    
    async def _read(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """读取工作表"""
        file_path = params["file_path"]
        sheet = params.get("sheet")
        cell_range = params.get("range")
        fmt = params.get("format", "json")
        has_header = params.get("has_header", True)
        
        if not os.path.exists(file_path):
            return {"success": False, "error": f"文件不存在: {file_path}"}
        
        sheet_arg = f'"{sheet}"' if sheet else "None"
        range_arg = f'"{cell_range}"' if cell_range else "None"
        
        cmd = [
            "python3", "-c",
            f"""
from openpyxl import load_workbook
import json

wb = load_workbook("{file_path}", read_only=True, data_only=True)
ws = wb[{sheet_arg}] if {sheet_arg} else wb.active

# 获取范围
if {range_arg}:
    cells = ws[{range_arg}]
else:
    cells = ws.iter_rows()

# 读取数据
rows = []
for row in cells:
    row_data = [cell.value for cell in row]
    if any(v is not None for v in row_data):  # 跳过空行
        rows.append(row_data)

# 处理格式
if {str(has_header).lower()} and rows:
    headers = rows[0]
    data = rows[1:]
else:
    headers = None
    data = rows

# 转换为JSON格式
if "{fmt}" == "json":
    if headers:
        result = [dict(zip(headers, row)) for row in data]
    else:
        result = data
elif "{fmt}" == "markdown":
    md = []
    if headers:
        md.append("| " + " | ".join(str(h) for h in headers) + " |")
        md.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in data:
        md.append("| " + " | ".join(str(v) if v is not None else "" for v in row) + " |")
    result = "\\n".join(md)
else:  # csv
    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    if headers:
        writer.writerow(headers)
    writer.writerows(data)
    result = output.getvalue()

print(json.dumps({{
    "sheet": ws.title,
    "rows_read": len(data),
    "data": result
}}, ensure_ascii=False))
            """
        ]
        
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )
        stdout, stderr = await proc.communicate()
        
        if proc.returncode != 0:
            return {"success": False, "error": stderr.decode()}
        
        result = json.loads(stdout.decode())
        return {"success": True, **result}
    
    async def _read_cell(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """读取单元格"""
        file_path = params["file_path"]
        cell = params["cell"]
        sheet = params.get("sheet")
        
        if not os.path.exists(file_path):
            return {"success": False, "error": f"文件不存在: {file_path}"}
        
        sheet_arg = f'"{sheet}"' if sheet else "None"
        
        cmd = [
            "python3", "-c",
            f"""
from openpyxl import load_workbook
import json

wb = load_workbook("{file_path}", read_only=True, data_only=True)
ws = wb[{sheet_arg}] if {sheet_arg} else wb.active

value = ws["{cell}"].value
result = {{
    "sheet": ws.title,
    "cell": "{cell}",
    "value": value,
    "type": type(value).__name__ if value is not None else None
}}
print(json.dumps(result, ensure_ascii=False))
            """
        ]
        
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )
        stdout, stderr = await proc.communicate()
        
        if proc.returncode != 0:
            return {"success": False, "error": stderr.decode()}
        
        result = json.loads(stdout.decode())
        return {"success": True, **result}
    
    async def _create(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """创建新工作簿"""
        output = params["output"]
        sheets = params.get("sheets", ["Sheet1"])
        
        sheets_str = str(sheets)
        
        cmd = [
            "python3", "-c",
            f"""
from openpyxl import Workbook
import json

wb = Workbook()
wb.remove(wb.active)  # 删除默认sheet

for sheet_name in {sheets_str}:
    wb.create_sheet(title=sheet_name)

wb.save("{output}")

result = {{
    "output": "{output}",
    "sheets_created": len({sheets_str})
}}
print(json.dumps(result, ensure_ascii=False))
            """
        ]
        
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )
        stdout, stderr = await proc.communicate()
        
        if proc.returncode != 0:
            return {"success": False, "error": stderr.decode()}
        
        result = json.loads(stdout.decode())
        return {"success": True, **result}
    
    async def _write(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """写入数据"""
        file_path = params["file_path"]
        data = params["data"]
        sheet = params.get("sheet")
        start_cell = params.get("start_cell", "A1")
        
        data_str = json.dumps(data)
        sheet_arg = f'"{sheet}"' if sheet else "None"
        
        cmd = [
            "python3", "-c",
            f"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

data = {data_str}
wb = load_workbook("{file_path}")
ws = wb[{sheet_arg}] if {sheet_arg} else wb.active

# 解析起始单元格
start_col = "{start_cell}"[0]
start_row = int("{start_cell}"[1:])

# 写入数据
for row_idx, row_data in enumerate(data):
    for col_idx, value in enumerate(row_data):
        col = get_column_letter(ord(start_col) - ord('A') + col_idx + 1)
        row = start_row + row_idx
        ws[f"{{col}}{{row}}"] = value

wb.save("{file_path}")

result = {{
    "file_path": "{file_path}",
    "rows_written": len(data),
    "columns": max(len(row) for row in data) if data else 0
}}
print(json.dumps(result, ensure_ascii=False))
            """
        ]
        
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )
        stdout, stderr = await proc.communicate()
        
        if proc.returncode != 0:
            return {"success": False, "error": stderr.decode()}
        
        result = json.loads(stdout.decode())
        return {"success": True, **result}
    
    async def _export(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """导出文件"""
        file_path = params["file_path"]
        output = params["output"]
        fmt = params["format"]
        sheet = params.get("sheet")
        
        if not os.path.exists(file_path):
            return {"success": False, "error": f"文件不存在: {file_path}"}
        
        sheet_arg = f'"{sheet}"' if sheet else "None"
        
        cmd = [
            "python3", "-c",
            f"""
from openpyxl import load_workbook
import json

wb = load_workbook("{file_path}", read_only=True, data_only=True)
ws = wb[{sheet_arg}] if {sheet_arg} else wb.active

# 读取所有数据
rows = []
for row in ws.iter_rows(values_only=True):
    if any(v is not None for v in row):
        rows.append(row)

if "{fmt}" == "csv":
    import csv
    with open("{output}", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
elif "{fmt}" == "json":
    if rows:
        headers = rows[0]
        data = [dict(zip(headers, row)) for row in rows[1:]]
    else:
        data = []
    with open("{output}", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
else:  # markdown
    md = []
    if rows:
        headers = rows[0]
        md.append("| " + " | ".join(str(h) for h in headers) + " |")
        md.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for row in rows[1:]:
            md.append("| " + " | ".join(str(v) if v is not None else "" for v in row) + " |")
    with open("{output}", "w", encoding="utf-8") as f:
        f.write("\\n".join(md))

result = {{
    "source": "{file_path}",
    "output": "{output}",
    "format": "{fmt}",
    "rows_exported": len(rows)
}}
print(json.dumps(result, ensure_ascii=False))
            """
        ]
        
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )
        stdout, stderr = await proc.communicate()
        
        if proc.returncode != 0:
            return {"success": False, "error": stderr.decode()}
        
        result = json.loads(stdout.decode())
        return {"success": True, **result}
